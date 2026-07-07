import requests
import feedparser
import calendar
import json
from bs4 import BeautifulSoup
from datetime import datetime, timezone, timedelta
from typing import Optional, Tuple, List, Dict
import re
import sys
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# ─────────────────────────────────────────────
# EDITE AQUI: fontes
# ─────────────────────────────────────────────

SOURCES = {
    "Valor Econômico": {
        "prefix": "V",
        "type": "sitemap_daily",
        # URLs estáticas — sempre buscadas (notícias recentes)
        "urls": [
            "https://valor.globo.com/sitemap/valor/news.xml",
        ],
        # Sitemaps diários — buscados conforme a janela de tempo
        # Formato: base/{ano}/{mes}/{dia}_{pagina}.xml
        # Datas organizadas em BRT (UTC-3); datas nas entradas têm offset -03:00 explícito
        "daily": {
            "format": "globo",
            "base": "https://valor.globo.com/sitemap/valor",
        },
    },
    "Estadão": {
        "prefix": "E",
        "type": "sitemap_daily",
        "urls": [
            "https://www.estadao.com.br/arc/outboundfeeds/sitemap/latest/?outputType=xml",
        ],
        # Formato: base/{ano}-{mes}-{dia}/?outputType=xml
        # Arc Publishing usa UTC (sufixo Z) nas datas — parse_iso converte automaticamente
        "daily": {
            "format": "estadao",
            "base": "https://www.estadao.com.br/arc/outboundfeeds/sitemap",
        },
    },
    "O Globo": {
        "prefix": "G",
        "type": "sitemap_daily",
        "urls": [
            "https://oglobo.globo.com/rss/oglobo",              # RSS padrão
            "https://oglobo.globo.com/sitemap/oglobo/news.xml", # notícias recentes
        ],
        # Formato igual ao Valor; datas em BRT com offset -03:00 explícito
        "daily": {
            "format": "globo",
            "base": "https://oglobo.globo.com/sitemap/oglobo",
        },
    },
    "Folha de São Paulo": {
        "prefix": "F",
        "type": "rss",
        "urls": [
            "https://feeds.folha.uol.com.br/mercado/rss091.xml",
            "https://feeds.folha.uol.com.br/poder/rss091.xml",
            "https://feeds.folha.uol.com.br/mundo/rss091.xml",
            "https://feeds.folha.uol.com.br/emcimadahora/rss091.xml",
            "https://feeds.folha.uol.com.br/ambiente/rss091.xml",
        ]
    },
}

# ─────────────────────────────────────────────
# Fim da área de configuração
# ─────────────────────────────────────────────

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
    ),
    "Accept-Language": "pt-BR,pt;q=0.9,en;q=0.8",
}
LOCAL_TZ = timezone(timedelta(hours=-3))


# ═══════════════════════════════════════════════════════════════════
# SISTEMA DE PONTUAÇÃO
# ═══════════════════════════════════════════════════════════════════

# ── Vocabulário permanente ─────────────────────────────────────────
# (frase_em_minúsculas, peso)
# Frases mais longas têm prioridade — não há double-count dos tokens.

_VOCAB_RAW: List[Tuple[str, int]] = [
    # 3.1 Política monetária e inflação
    ("decisão do copom",          12),
    ("comunicado do copom",        12),
    ("reunião do copom",           12),
    ("taxa selic",                 11),
    ("meta de inflação",           11),
    ("política monetária",         11),
    ("ciclo de alta de juros",     10),
    ("ciclo de queda de juros",    10),
    ("ciclo de cortes",            10),
    ("convergência da inflação",   10),
    ("expectativas de inflação",   10),
    ("expectativa de inflação",    10),
    ("copom",                      10),
    ("juros futuros",               9),
    ("curva de juros",              9),
    ("banco central",               9),
    ("selic",                       9),
    ("inflação",                    9),
    ("ipca",                        9),
    ("federal reserve",             8),
    ("corte de juros",              8),
    ("redução de juros",            8),
    ("reduz juros",                 8),
    ("alta de juros",               8),
    ("elevação de juros",           8),
    ("eleva juros",                 8),
    ("aumenta juros",               8),
    ("diminui juros",               8),
    ("reservas internacionais",     8),
    ("desdolarização",              8),
    ("projeção de inflação",        8),
    ("projeção de juros",           8),
    ("projeta inflação",            8),
    ("projeta juros",               8),
    ("taxa de juros",               7),
    ("reservas de ouro",            7),
    ("expectativa de juros",        7),
    ("juros",                       7),
    ("dólar",                       7),
    ("câmbio",                      7),
    ("banco central europeu",       5),
    ("banco do japão",              5),
    # 3.2 Fiscal
    ("arcabouço fiscal",           11),
    ("meta fiscal",                11),
    ("contas públicas",            11),
    ("resultado primário",         11),
    ("déficit fiscal",             10),
    ("dívida pública",             10),
    ("pautas-bomba",               10),
    ("política fiscal",            10),
    ("orçamento público",           9),
    ("gastos públicos",             9),
    ("gasto público",               9),
    ("despesas obrigatórias",       9),
    ("renúncia fiscal",             9),
    ("risco fiscal",                9),
    ("tesouro nacional",            8),
    ("emissão de títulos",          8),
    ("agências de risco",           8),
    ("agência de risco",            8),
    ("risco soberano",              8),
    ("precatórios",                 8),
    ("reforma tributária",          8),
    ("reforma trabalhista",         8),
    ("bloqueio orçamentário",       6),
    ("orçamento",                   7),
    ("estatais",                    7),
    ("subsídios",                   7),
    ("subsídio",                    7),
    ("tributação",                  7),
    ("impostos",                    7),
    ("contingenciamento",           6),
    ("arrecadação",                 6),
    ("gastos",                      6),
    ("despesas",                    6),
    ("receitas",                    6),
    ("salário mínimo",              8),
    ("previdência",                 5),
    ("benefícios sociais",          5),
    # 3.3 Mercados financeiros
    ("risco-país",                 10),
    ("crise financeira",           10),
    ("crise bancária",             10),
    ("ibovespa",                    9),
    ("títulos públicos",            9),
    ("mercado financeiro",          8),
    ("bolsas asiáticas",            8),
    ("bolsas globais",              8),
    ("brent",                       8),
    ("wti",                         8),
    ("títulos brasileiros",         8),
    ("petróleo",                    9),
    ("prêmio de risco",             7),
    ("renda fixa",                  7),
    ("bolsa",                       7),
    ("bolsas",                      7),
    ("commodities",                 7),
    ("ouro",                        7),
    ("ações",                       4),
    ("dividendos",                  6),
    ("balanço",                     6),
    ("lucro",                       6),
    ("prejuízo",                    6),
    ("fusão",                       6),
    ("aquisição",                   6),
    ("ipo",                         6),
    # 3.4 Bancos, crédito e sistema financeiro
    ("crédito consignado",          8),
    ("ativos virtuais",             8),
    ("regulação bancária",          8),
    ("sistema financeiro",          8),
    ("liquidação de banco",         8),
    ("intervenção em banco",        8),
    ("bancos",                      8),
    ("consignado",                  7),
    ("fintech",                     7),
    ("crédito",                     7),
    ("inadimplência",               6),
    ("endividamento",               6),
    ("dívida do consumidor",        6),
    ("crédito privado",             6),
    # 3.5 Energia, infraestrutura e trabalho
    ("tarifa de energia",          10),
    ("tarifas de energia",         10),
    ("conta de luz",               10),
    ("reforma da previdência",     10),
    ("jornada 6x1",                 9),
    ("escala 6x1",                  9),
    ("fim da jornada 6x1",          9),
    ("fim da escala 6x1",           9),
    ("mercado de trabalho",         8),
    ("pejotização",                 8),
    ("emprego",                     8),
    ("desemprego",                  8),
    ("salários",                    8),
    ("setor elétrico",              6),
    ("geração de energia",          6),
    ("energia",                     6),
    ("hidrelétricas",               6),
    ("termelétricas",               6),
    ("térmicas",                    6),
    ("gás natural",                 6),
    ("eólicas offshore",            6),
    ("infraestrutura",              6),
    ("transportes",                 6),
    ("tarifa de ônibus",            6),
    # 3.6 Comércio exterior e geopolítica econômica
    ("guerra comercial",           10),
    ("sanções econômicas",         10),
    ("bloqueio econômico",         10),
    ("tarifaço",                    9),
    ("barreira comercial",          9),
    ("restrição de exportações",    9),
    ("comércio exterior",           8),
    ("retaliação comercial",        8),
    ("acordo comercial",            8),
    ("estreito de ormuz",           8),
    ("ormuz",                       8),
    ("exportações",                 7),
    ("importações",                 7),
    ("tarifas",                     7),
    ("sanções",                     7),
    ("retaliação",                  7),
    ("yuan",                        7),
    ("moeda chinesa",               7),
    ("mercosul",                    6),
    ("guerra",                      3),
    ("conflito",                    3),
    ("acordo",                      3),
    ("negociação",                  3),
    ("relações comerciais",         3),
    # 3.7 Empresas e atividade econômica
    ("petrobras",                   9),
    ("vale",                        9),
    ("eletrobras",                  9),
    ("bndes",                       9),
    ("indústria",                   9),
    ("varejo",                      9),
    ("serviços",                    9),
    ("agronegócio",                 9),
    ("investimento",                6),
    ("produção",                    6),
    ("exportação",                  6),
    ("endividamento corporativo",   6),
    ("construção",                  5),
    ("mineração",                   5),
    ("setor automotivo",            5),
    ("crescimento econômico",       4),
    ("atividade econômica",         4),
    ("produtividade",               4),
    ("competitividade",             4),
    # 3.8 Instituições e regulação
    ("decisão do stf",              6),
    ("regulação",                   6),
    ("agências reguladoras",        6),
    ("agência reguladora",          6),
    ("supremo",                     5),
    ("congresso",                   5),
    ("senado",                      5),
    ("câmara",                      5),
    ("medida provisória",           4),
    ("projeto de lei",              4),
    ("súmula",                      4),
    ("julgamento",                  4),
    ("governo",                     3),
    ("ministério",                  3),
    ("presidente",                  3),
    ("relator",                     3),
    # 3.9 Política eleitoral
    ("pesquisa eleitoral",         10),
    ("datafolha",                  10),
    ("intenção de voto",           10),
    ("primeiro turno",             10),
    ("segundo turno",              10),
    ("eleições",                    7),
    ("candidato",                   7),
    ("candidatura",                 7),
    ("reeleição",                   7),
    ("campanha",                    4),
    ("pré-campanha",                4),
    ("base eleitoral",              4),
]

# Siglas que requerem correspondência de palavra inteira (case-sensitive no original)
SIGLAS_EXATAS = {
    "BC": 8, "PF": 7, "PGR": 7, "STF": 7, "TSE": 5, "TCU": 7,
    "UE": 4, "EUA": 4, "Fed": 8, "BCE": 5, "BoJ": 5,
    "CCJ": 5, "PL": 4, "PEC": 6, "MP": 4, "CVM": 7,
    "G7": 6, "G20": 6, "Opep": 6, "OPEC": 6, "OMC": 6, "FMI": 6, "BID": 6,
    "BRB": 7, "FGTS": 7, "IPCA": 9, "PIB": 4, "CDI": 6,
    "BNDES": 9,
}

# Termos contextuais — só pontuam com condição
CONTEXTUAIS = {
    "economistas": {"peso": 5, "requer_min": 7},    # só com termo ≥7
    "mercado":     {"peso": 3, "requer_min": 5},    # só com termo ≥5
    "empresa":     {"peso": 3, "requer_min": 5},
    "empresas":    {"peso": 3, "requer_min": 5},
    "setor":       {"peso": 3, "requer_min": 5},
    # Países — só com contexto econômico
    "china":       {"peso": 4, "requer_min": 5},
    "irã":         {"peso": 4, "requer_min": 5},
    "rússia":      {"peso": 4, "requer_min": 5},
    "japão":       {"peso": 4, "requer_min": 5},
    "europa":      {"peso": 4, "requer_min": 5},
    # Bets — só com dívida/crédito
    "bets":        {"peso": 4, "requer_palavras": {"dívida","endividamento","crédito","renda","contas","inadimplência"}},
    "apostas":     {"peso": 4, "requer_palavras": {"dívida","endividamento","crédito","renda","contas","inadimplência"}},
    # Clima — só com energia/alimentos
    "el niño":     {"peso": 5, "requer_palavras": {"energia","agricultura","inflação","alimentos","tarifa","hidrelétrica","conta de luz"}},
    "la niña":     {"peso": 5, "requer_palavras": {"energia","agricultura","inflação","alimentos","tarifa","hidrelétrica","conta de luz"}},
    "seca":        {"peso": 5, "requer_palavras": {"energia","agricultura","inflação","alimentos","tarifa","hidrelétrica","conta de luz"}},
    "chuvas":      {"peso": 5, "requer_palavras": {"energia","agricultura","inflação","alimentos","tarifa","hidrelétrica","conta de luz"}},
    "clima":       {"peso": 5, "requer_palavras": {"energia","agricultura","inflação","alimentos","tarifa","hidrelétrica","conta de luz"}},
}

# Nomes próprios: não pontuam no vocab mas PODEM formar clusters e tendências
VOCAB_ZERO: set = set()

# Termos que NUNCA formam cluster
CLUSTER_BLOCK = {"openai", "spacex"}

# Compatibilidade retroativa — usado no scoring legado
PESO_ZERO = VOCAB_ZERO

# Países: só viram cluster se ≥2 aparições forem de editorias econômico-políticas
COUNTRY_NAMES = {
    "argentina","chile","méxico","colômbia","venezuela","peru","bolívia","uruguai",
    "paraguai","cuba","equador","haiti","nicarágua","panamá","costa rica","eua",
    "alemanha","frança","reino unido","itália","espanha","portugal","rússia",
    "ucrânia","polônia","hungria","suíça","holanda","china","japão","índia",
    "coreia","taiwan","israel","irã","turquia","arábia saudita","vietnã",
    "singapura","austrália","canadá","áfrica do sul","egito","nigéria",
}
_COUNTRY_VALID_EDITORIAS = {
    "economia","economias","financas","politica","poder","mercado","mercados",
    "agronegocio","legislacao","negocios","einvestidor",
}

# Organiza vocab para matching phrase-first (mais longo primeiro)
VOCAB_SORTED = sorted(_VOCAB_RAW, key=lambda x: len(x[0]), reverse=True)


# ── Ações / acontecimentos ─────────────────────────────────────────

# Grupo → (verbos/frases, bônus)
ACOES_GRUPOS = {
    "decisão_econômica": (
        {"corta","cortou","reduz","reduziu","eleva","elevou",
         "aumenta","aumentou","mantém","manteve","decide","decidiu",
         "define","definiu"},
        5,
        # requer objeto econômico
        {"juros","selic","taxa","inflação","tarifa","preço","preços",
         "produção","gastos","dívida","crédito","exportações","câmbio","dólar","orçamento"},
    ),
    "decisão_legislativa": (
        {"aprova","aprovou","rejeita","rejeitou","sanciona","sancionou",
         "veta","vetou","suspende","suspendeu","barra","barrou",
         "autoriza","autorizou","condena","condenou","julga","julgou",
         "endurece","endureceu"},
        5, None,
    ),
    "investigação_grave": (
        {"cumpre buscas","operação da pf","deflagra operação","propina",
         "desvio","desvios","caixa paralelo","delação premiada","delação",
         "indicia","indiciou","denuncia","denunciou","fraude","corrupção",
         "lavagem de dinheiro","prisão"},
        6, None,
    ),
    "investigação_menor": (
        {"investiga","investigação","mira","aponta pf","segundo a pf","relatório da pf"},
        4, None,
    ),
    "medida": (
        {"anuncia","anunciou","lança","lançou","emite","emissão",
         "restringe","restringiu","adia","adiou","extingue","cria","amplia"},
        3, None,
    ),
    "possibilidade": (
        {"estuda","avalia","discute","cogita","pode","deve","espera",
         "prevê","aposta","considera"},
        3, None,
    ),
}

# ── Impacto ────────────────────────────────────────────────────────

IMPACTO_FRASES: List[Tuple[str, int]] = [
    ("pressiona inflação", 4), ("pressiona juros", 4),
    ("encarece tarifas", 4), ("encarece energia", 4),
    ("eleva gastos", 4), ("aumenta dívida", 4),
    ("reduz inflação", 4), ("derruba juros", 4),
    ("restringe exportações", 4), ("afeta contas públicas", 4),
    ("impacta", 3), ("afeta", 3), ("ameaça", 3), ("encarece", 3),
    ("pressiona", 3), ("alivia", 3), ("despenca", 3), ("dispara", 3),
    ("turbulência", 3), ("risco", 3),
    ("sobe", 2), ("cai", 2), ("avança", 2), ("recua", 2),
    ("alta", 2), ("queda", 2), ("aumento", 2), ("redução", 2),
    ("pode afetar", 1), ("pode elevar", 1),
    ("pode reduzir", 1), ("pode pressionar", 1),
]
IMPACTO_SORTED = sorted(IMPACTO_FRASES, key=lambda x: len(x[0]), reverse=True)

# ── Penalidades ────────────────────────────────────────────────────

PENALIDADES_LIST: List[Tuple[str, int]] = [
    # Esportes
    ("futebol", -35), ("campeonato", -35), ("gol", -35),
    ("time", -35), ("jogador", -35), ("técnico", -35), ("torcida", -35),
    ("esporte", -35), ("esportes", -35),
    # Entretenimento
    ("bbb", -30), ("reality show", -30), ("celebridade", -30),
    ("famoso", -30), ("famosa", -30), ("novela", -30), ("horóscopo", -30),
    # Cultura sem impacto
    ("música", -25), ("cinema", -25), ("série", -25),
    ("show", -25), ("entretenimento", -25), ("moda", -25), ("beleza", -25),
    # Gastronomia
    ("receita", -20), ("culinária", -20), ("gastronomia", -20), ("restaurante", -20),
    # Turismo
    ("turismo", -15), ("viagem", -15), ("hotel", -15), ("praia", -15), ("destino", -15),
    # Opinião
    ("artigo de opinião", -20), ("editorial", -20), ("opinião", -20),
    ("crônica", -12), ("ensaio", -12), ("resenha", -12),
    # Crime local sem relevância econômica
    ("assassinato", -20), ("homicídio", -20), ("roubo", -20),
    ("assalto", -20), ("acidente", -20), ("desaparecimento", -20),
    # Clickbait
    ("veja", -3), ("saiba", -3), ("entenda", -3), ("confira", -3), ("descubra", -3),
    # Loteria
    ("lotofácil", -30), ("mega-sena", -30), ("número sorteado", -20),
    # Podcast
    ("coluna de podcast", -20), ("podcast", -20),
]
PENALIDADES_SORTED = sorted(PENALIDADES_LIST, key=lambda x: len(x[0]), reverse=True)

# Penalidades que só aplicam se a editoria for neutra/negativa (≤ 0 pts)
# Palavras ambíguas: "jogo" pode ser game regulatório; "filme" pode ser caso no STF
PENALIDADES_CONDICIONAIS: List[Tuple[str, int]] = [
    ("jogo",  -35),
    ("filme", -25),
]
PENALIDADES_COND_SORTED = sorted(PENALIDADES_CONDICIONAIS, key=lambda x: len(x[0]), reverse=True)



# Palavras a ignorar na extração automática de entidades
_STOP_ENTITIES = {
    "para","uma","uns","das","dos","com","por","que","não","mais",
    "mas","como","são","pela","pelo","seu","sua","seus","suas",
    "isso","este","esta","esse","essa","esses","essas","estes","estas",
    "após","ante","até","desde","entre","sobre","contra","além","novo",
    "nova","novos","novas","grande","grande","primeiro","segunda","anos",
    "meses","dias","horas","brasil","governo","ministro","ministros",
    "presidente",
    # Frases multi-palavra que aparecem como entidades falsas
    "diz que","ibovespa hoje","roda viva","são paulo","multifranqueados","figurinhas",
}

# Termos que NUNCA podem ser tendências
TECHNICAL_NOISE = {
    "ghtml", "html", "http", "https", "www",
    "nprec", "npres", "nprp",
    "globo", "estadao", "folha", "valor",
}

CLICKBAIT_WORDS = {
    "veja", "entenda", "saiba", "confira", "descubra",
    "quem", "quando", "onde", "quais", "como", "por que",
    "hoje", "agora",
}

GENERIC_WORDS = {
    "mundo", "estado", "empresa", "empresas", "grupo",
    "mercado", "mercados", "venda", "compra", "queda", "alta",
    "aumento", "redução", "projeto", "conselho", "semana",
    "história", "futuro", "onda", "caso", "crise", "taxa",
    "pressão", "disputa", "análise", "brasileiro", "mulher",
    # Números e quantidades
    "dois","três","quatro","cinco","dez","cem","mil",
    "milhões","bilhões","reais","porcentagem",
    # Tempo genérico
    "meses","semanas","terça","quarta","quinta","sexta",
    # Do PDF — marcados como não-tema
    "aliados","custo","valores","motoristas","quebra",
    "copa","brics","pasep","benefício","benefícios",
    # Loteria
    "lotofácil","mega-sena","número sorteado",
    # Outros falsos positivos comuns
    "resultado","resultados","reunião","aprovação","proposta",
    "acesso","medida","plano","programa","resposta","carta",
    "saída","volta","frente","morte","mortes","lista",
    "nota","falta","parte","vez","vida","acordo",
}

# União de todos os termos proibidos como tendência
_TREND_STOPWORDS = _STOP_ENTITIES | TECHNICAL_NOISE | CLICKBAIT_WORDS | GENERIC_WORDS

# Tópicos permitidos para análise de tendências, por veículo
ALLOWED_TOPICS: Dict[str, set] = {
    "Folha de São Paulo": {"mercado", "poder", "ambiente", "mundo"},
    "Valor Econômico":    {"financas", "politica", "agronegocio", "legislacao", "mundo"},
    "Estadão":            {"economia", "politica", "einvestidor", "internacional"},
    "O Globo":            {"economia", "politica", "mundo"},
}


def _article_topic(url: str) -> Optional[str]:
    """Extrai o tópico/editoria da URL do artigo."""
    m = re.search(
        r'(?:valor\.globo\.com|oglobo\.globo\.com|estadao\.com\.br|folha\.uol\.com\.br)'
        r'/([a-z][a-z-]+)/',
        url.lower(),
    )
    return m.group(1) if m else None


def _topic_allowed(article: dict) -> bool:
    """True se o artigo pertence a um tópico permitido para análise de tendências."""
    source = article.get("source", "")
    allowed = ALLOWED_TOPICS.get(source)
    if allowed is None:
        return True  # fonte não mapeada: permite
    topic = _article_topic(article.get("url", ""))
    return topic in allowed if topic else False


# ═══════════════════════════════════════════════════════════════════
# MOTOR DE PONTUAÇÃO
# ═══════════════════════════════════════════════════════════════════

def _token_boundary(phrase: str, text: str) -> bool:
    """Verifica se 'phrase' aparece com limites de palavra em 'text' (ambos já em lower)."""
    pattern = r"(?<!\w)" + re.escape(phrase) + r"(?!\w)"
    return bool(re.search(pattern, text))


def _sigla_match(sigla: str, original: str) -> bool:
    """Match exato de sigla preservando case no texto original."""
    pattern = r"(?<!\w)" + re.escape(sigla) + r"(?!\w)"
    return bool(re.search(pattern, original))


def _match_vocab_permanente(title_lower: str, title_orig: str
                             ) -> Tuple[int, List[str], int]:
    """
    Retorna (score, explicações, max_weight_encontrado).
    Usa phrase-first: tokens de uma frase já contabilizada não pontuam de novo.
    """
    claimed: set = set()
    score = 0
    explanations = []
    max_w = 0

    # Siglas exatas (case-sensitive)
    for sigla, peso in SIGLAS_EXATAS.items():
        if _sigla_match(sigla, title_orig):
            words = set(sigla.lower().split())
            if not words & claimed:
                claimed |= words
                score += peso
                max_w = max(max_w, peso)
                explanations.append(f"{sigla} +{peso}")

    # Frases do vocabulário (phrase-first, mais longo primeiro)
    for phrase, peso in VOCAB_SORTED:
        if _token_boundary(phrase, title_lower):
            words = set(phrase.split())
            if not words & claimed:
                claimed |= words
                score += peso
                max_w = max(max_w, peso)
                explanations.append(f"{phrase} +{peso}")

    # Termos contextuais
    for termo, cfg in CONTEXTUAIS.items():
        if not _token_boundary(termo, title_lower):
            continue
        if set(termo.split()) & claimed:
            continue
        if "requer_min" in cfg:
            if max_w >= cfg["requer_min"]:
                score += cfg["peso"]
                claimed |= set(termo.split())
                explanations.append(f"{termo} (contextual) +{cfg['peso']}")
        elif "requer_palavras" in cfg:
            if any(_token_boundary(req, title_lower) for req in cfg["requer_palavras"]):
                score += cfg["peso"]
                claimed |= set(termo.split())
                explanations.append(f"{termo} (contextual) +{cfg['peso']}")

    return score, explanations, max_w


def _match_acoes(title_lower: str, base_score: int) -> Tuple[int, str]:
    """Retorna (bônus, explicação). Aplica apenas o maior bônus encontrado."""
    best = (0, "")
    for grupo, (verbos, bonus, requer_obj) in ACOES_GRUPOS.items():
        for v in verbos:
            if not _token_boundary(v, title_lower):
                continue
            if requer_obj and not any(_token_boundary(o, title_lower) for o in requer_obj):
                continue
            if bonus > best[0]:
                best = (bonus, f"{v} ({grupo.replace('_',' ')}) +{bonus}")
            break
    return best


def _match_impacto(title_lower: str, base_score: int) -> Tuple[int, List[str]]:
    """Retorna (total impacto, explicações). Máximo +8."""
    if base_score < 3:
        return 0, []
    claimed: set = set()
    total = 0
    exps = []
    for phrase, peso in IMPACTO_SORTED:
        if total >= 8:
            break
        if _token_boundary(phrase, title_lower):
            words = set(phrase.split())
            if not words & claimed:
                add = min(peso, 8 - total)
                claimed |= words
                total += add
                exps.append(f"impacto '{phrase}' +{add}")
    return total, exps


def _match_especificidade(title_orig: str, base_score: int) -> Tuple[int, List[str]]:
    """Retorna (bônus especificidade, explicações). Máximo +5. Requer base ≥5."""
    if base_score < 5:
        return 0, []
    total = 0
    exps = []
    # +3 número (percentual, R$, US$, quantidade)
    if re.search(r'\d+[,\.]\d+\s*%|\bR\$\s*[\d,\.]+|\bUS\$\s*[\d,\.]+|\b\d+\s*(milhões?|bilhões?|trilhões?)\b', title_orig, re.IGNORECASE):
        total += 3
        exps.append("valor/percentual numérico +3")
    # +2 instituição + ação + objeto (heurística: título curto com 3+ tokens pontuados)
    words = title_orig.split()
    if len(words) <= 12 and base_score >= 10:
        total = min(total + 2, 5)
        exps.append("manchete objetiva (inst+ação+objeto) +2")
    # +1 data futura ou mês/ano
    if re.search(r'\b(janeiro|fevereiro|março|abril|maio|junho|julho|agosto|setembro|outubro|novembro|dezembro|20\d\d)\b', title_orig, re.IGNORECASE):
        remaining = 5 - total
        if remaining > 0:
            total += 1
            exps.append("prazo/data específica +1")
    return min(total, 5), exps


def _match_penalidades(title_lower: str, subtotal_positivo: int = 0,
                       has_cluster: bool = False) -> Tuple[int, List[str]]:
    total = 0
    exps = []
    claimed: set = set()
    for phrase, peso in PENALIDADES_SORTED:
        if _token_boundary(phrase, title_lower):
            words = set(phrase.split())
            if not words & claimed:
                claimed |= words
                total += peso
                exps.append(f"'{phrase}' {peso}")
    # Penalidades condicionais: ignoradas se artigo já tem sinal relevante de outros fatores
    # ou se alguma palavra do título está num cluster ativo
    skip_cond = subtotal_positivo >= 7 or has_cluster
    if not skip_cond:
        for phrase, peso in PENALIDADES_COND_SORTED:
            if _token_boundary(phrase, title_lower):
                words = set(phrase.split())
                if not words & claimed:
                    claimed |= words
                    total += peso
                    exps.append(f"'{phrase}' {peso} (cond.)")
    return total, exps



def _calc_editoria(url: str) -> Tuple[int, str]:
    u = url.lower()
    for seg, pts, label in [
        # Positivos
        ("/economia/",      3,   "editoria economia"),
        ("/financas/",      3,   "editoria finanças"),
        ("/mercados/",      3,   "editoria mercados"),
        ("/politica/",      3,   "editoria política"),
        ("/empresas/",      1,   "editoria empresas"),
        ("/internacional/", 1,   "editoria internacional"),
        ("/mundo/",         1,   "editoria mundo"),
        ("/einvestidor/",   3,   "editoria einvestidor"),
        ("/mercado/",       3,   "editoria mercado"),
        ("/poder/",         3,   "editoria poder"),
        # Negativos
        ("/esportes/",      -10, "editoria esportes"),
        ("/cultura/",       -10, "editoria cultura"),
        ("/podcasts/",      -10, "editoria podcasts"),
        ("/viagem/",        -10, "editoria viagem"),
        ("/paladar/",       -10, "editoria paladar"),
        ("/boa-viagem/",    -10, "editoria boa-viagem"),
    ]:
        if seg in u:
            sign = "+" if pts > 0 else ""
            return pts, f"{label} {sign}{pts}"
    return 0, ""


# ── Clusters e tendências ──────────────────────────────────────────

def _extract_entities(title: str) -> set:
    """Extrai entidades nomeadas (palavras capitalizadas não no vocab nem listas de exclusão).
    VOCAB_ZERO é excluído do vocab mas permitido como entidade para formar clusters."""
    known_lower = {p for p, _ in _VOCAB_RAW} | set(SIGLAS_EXATAS.keys())
    pattern = r'\b[A-ZÁÉÍÓÚÂÊÎÔÛÃÕ][a-záéíóúâêîôûãõA-ZÁÉÍÓÚÂÊÎÔÛÃÕ]{2,}(?:\s+[A-ZÁÉÍÓÚÂÊÎÔÛÃÕ][a-záéíóúâêîôûãõA-ZÁÉÍÓÚÂÊÎÔÛÃÕ]{2,})?\b'
    found = set()
    for m in re.findall(pattern, title):
        ml = m.lower()
        if ml not in known_lower and ml not in _TREND_STOPWORDS and ml not in CLUSTER_BLOCK and len(m) > 3:
            found.add(m)
    return found


def _build_alias_map(articles: List[dict]) -> Dict[str, str]:
    """
    Mapeia componentes únicos de nomes compostos para o nome canônico.
    Ex: se 'Jaques Wagner' foi visto e 'Wagner' não aparece em outros nomes compostos,
    então alias_map['wagner'] = 'jaques wagner'.
    """
    multi_word_seen: set = set()
    for art in articles:
        for ent in _extract_entities(art["title"]):
            if " " in ent:
                multi_word_seen.add(ent.lower())

    # Para cada componente, rastrear em quantos nomes compostos aparece
    component_to_mw: Dict[str, set] = {}
    for mw in multi_word_seen:
        for part in mw.split():
            if len(part) > 3 and part not in _TREND_STOPWORDS:
                component_to_mw.setdefault(part, set()).add(mw)

    # Alias só se componente é inequívoco (mapeado para um único nome composto)
    alias_map: Dict[str, str] = {}
    for component, mw_set in component_to_mw.items():
        if len(mw_set) == 1:
            alias_map[component] = next(iter(mw_set))

    return alias_map


def _count_entities(articles: List[dict], alias_map: Dict[str, str]) -> Dict[str, dict]:
    """Conta entidades por artigo, resolvendo aliases para o nome canônico."""
    counter: Dict[str, dict] = {}
    for art in articles:
        if not _topic_allowed(art):
            continue
        for ent in _extract_entities(art["title"]):
            ek = ent.lower()
            canonical = alias_map.get(ek, ek)
            if canonical not in counter:
                # Prefere a forma mais longa como display
                counter[canonical] = {"count": 0, "sources": set(),
                                       "display": ent, "terms": set()}
            elif len(ent) > len(counter[canonical]["display"]):
                counter[canonical]["display"] = ent
            counter[canonical]["count"] += 1
            counter[canonical]["sources"].add(art.get("source", "?"))
            counter[canonical]["terms"].add(ent)
    return counter


def build_cluster_stats(
    current_articles: List[dict],
    historical_articles: Optional[List[dict]] = None,
) -> Dict[str, dict]:
    """
    Constrói clusters com:
    - Alias por componente (Jaques + Wagner → Jaques Wagner)
    - Merge por co-ocorrência (Vorcaro + Banco Master → cluster único)
    - Validação de países por editoria
    - Amplificação histórica com artigos fora da janela
    """
    # === 1. Alias map (baseado em todos os artigos disponíveis) ===
    all_arts = current_articles + (historical_articles or [])
    alias_map = _build_alias_map(all_arts)

    # === 2. Contagem separada: janela atual vs histórico ===
    curr_counter = _count_entities(current_articles, alias_map)
    hist_counter = _count_entities(historical_articles or [], alias_map)

    # === 3. Co-ocorrência nos artigos atuais → Union-Find ===
    # Monta artigos → entidades canônicas
    art_ents: List[set] = []
    for art in current_articles:
        if not _topic_allowed(art):
            continue
        cset = set()
        for ent in _extract_entities(art["title"]):
            cset.add(alias_map.get(ent.lower(), ent.lower()))
        art_ents.append(cset)

    cooccur: Dict[tuple, int] = {}
    for ents in art_ents:
        eligible = [e for e in ents if curr_counter.get(e, {}).get("count", 0) >= 2
                    and e not in CLUSTER_BLOCK]
        for i in range(len(eligible)):
            for j in range(i + 1, len(eligible)):
                pair = tuple(sorted([eligible[i], eligible[j]]))
                cooccur[pair] = cooccur.get(pair, 0) + 1

    # Union-Find
    _parent: Dict[str, str] = {}
    def _find(x: str) -> str:
        _parent.setdefault(x, x)
        if _parent[x] != x:
            _parent[x] = _find(_parent[x])
        return _parent[x]
    def _union(a: str, b: str) -> None:
        ra, rb = _find(a), _find(b)
        if ra == rb:
            return
        ca = curr_counter.get(ra, {}).get("count", 0)
        cb = curr_counter.get(rb, {}).get("count", 0)
        if ca >= cb:
            _parent[rb] = ra
        else:
            _parent[ra] = rb

    for (a, b), co in cooccur.items():
        fa = curr_counter.get(a, {}).get("count", 0)
        fb = curr_counter.get(b, {}).get("count", 0)
        if fa < 2 or fb < 2:
            continue
        ratio = co / min(fa, fb)
        if ratio >= 0.65 and co >= 2:
            _union(a, b)

    # Agrupa membros por raiz; limita a 4 por grupo
    root_groups: Dict[str, set] = {}
    qualifying = {ek for ek, info in curr_counter.items() if info["count"] >= 2}
    for ek in qualifying:
        root = _find(ek)
        root_groups.setdefault(root, set()).add(ek)

    for root, members in list(root_groups.items()):
        if len(members) > 4:
            top4 = sorted(members, key=lambda x: curr_counter.get(x, {}).get("count", 0), reverse=True)[:4]
            for m in members:
                if m not in top4:
                    _parent[m] = m  # desacopla do grupo
            root_groups[root] = set(top4)

    # === 4. Validação de países por editoria ===
    def _country_valid(canonical: str) -> bool:
        if canonical not in COUNTRY_NAMES:
            return True
        valid = 0
        for art in current_articles:
            if not _topic_allowed(art):
                continue
            topic = _article_topic(art.get("url", "")) or ""
            if topic.lower() in _COUNTRY_VALID_EDITORIAS:
                for ent in _extract_entities(art["title"]):
                    if alias_map.get(ent.lower(), ent.lower()) == canonical:
                        valid += 1
                        break
        return valid >= 2

    # === 5. Monta stats finais ===
    stats: Dict[str, dict] = {}

    processed_roots = set()
    for ek in list(qualifying):
        root = _find(ek)
        if root in processed_roots:
            continue
        processed_roots.add(root)
        members = root_groups.get(root, {root})

        # Descarta países sem aparições em editorias válidas
        if not all(_country_valid(m) for m in members):
            continue

        is_grouped = len(members) > 1

        # Stats janela atual
        c_count   = sum(curr_counter.get(m, {}).get("count", 0) for m in members)
        c_sources = set().union(*(curr_counter.get(m, {}).get("sources", set()) for m in members))
        all_terms = set().union(*(curr_counter.get(m, {}).get("terms", set()) for m in members))

        # Stats histórico (só considera se count ≥ 3)
        h_raw   = sum(hist_counter.get(m, {}).get("count", 0) for m in members)
        h_count = h_raw if h_raw >= 3 else 0
        h_sources = (set().union(*(hist_counter.get(m, {}).get("sources", set()) for m in members))
                     if h_count > 0 else set())

        all_sources = c_sources | h_sources
        eff_count = c_count + h_count * 0.4

        # Pontuação: eff direto, limitado pelo cap
        cap   = 8 if is_grouped else 7
        bonus = min(int(eff_count), cap)

        # Nome canônico = membro com maior count
        best = max(members, key=lambda m: curr_counter.get(m, {}).get("count", 0))
        display = curr_counter[best]["display"] if best in curr_counter else best.title()

        # Explicação legível
        hist_note = f" + {h_raw}×0.4={h_count*0.4:.1f}hist" if h_count > 0 else ""
        exp = (f"janela={c_count}{hist_note} → eff={eff_count:.1f} | "
               f"fontes={len(c_sources)} jan")
        if h_sources - c_sources:
            exp += f"+{len(h_sources - c_sources)} hist"
        if is_grouped:
            exp += f" | grupo: {', '.join(sorted(members))}"

        stats[display] = {
            "count":              round(eff_count, 1),
            "count_current":      c_count,
            "count_historical":   h_count,
            "sources":            all_sources,
            "sources_current":    c_sources,
            "sources_historical": h_sources,
            "terms_found":        all_terms | {display},
            "is_grouped":         is_grouped,
            "group_members":      sorted(members),
            "bonus":              bonus,
            "bonus_explanation":  exp,
        }

    # === 6. Entidades apenas no histórico (sem aparecimento na janela atual) ===
    for ek, hinfo in hist_counter.items():
        if hinfo["count"] < 3:
            continue
        # Já coberta por merge acima?
        if any(ek in st.get("group_members", []) or ek == st["bonus_explanation"]
               for st in stats.values()):
            continue
        # Verifica se o display já está em stats
        disp = hinfo["display"]
        if any(disp.lower() == k.lower() for k in stats):
            continue
        h_count  = hinfo["count"]
        h_src    = hinfo["sources"]
        eff      = h_count * 0.4
        bonus = min(int(eff), 4)  # cap 4 para histórico puro
        exp = f"apenas histórico: {h_count}×0.4={eff:.1f}eff | fontes={len(h_src)}"
        stats[disp] = {
            "count":              round(eff, 1),
            "count_current":      0,
            "count_historical":   h_count,
            "sources":            h_src,
            "sources_current":    set(),
            "sources_historical": h_src,
            "terms_found":        hinfo.get("terms", {disp}),
            "is_grouped":         False,
            "group_members":      [],
            "bonus":              bonus,
            "bonus_explanation":  exp,
        }

    return stats


def _tendencia_bonus(title_lower: str, title_orig: str,
                     cluster_stats: Dict[str, dict]) -> Tuple[int, List[str]]:
    """Retorna (bônus total tendência, explicações). Máximo +8."""
    exps = []
    total = 0

    for cname, st in cluster_stats.items():
        terms = st["terms_found"]
        hits = [t for t in terms if _token_boundary(t.lower(), title_lower)]
        if not hits:
            continue

        bonus = st["bonus"]
        if bonus <= 0:
            continue

        remaining = 8 - total
        actual = min(bonus, remaining)
        if actual > 0:
            total += actual
            exps.append(f"tendência '{cname}' ({st['bonus_explanation']}) +{actual}")
            if total >= 8:
                break

    return total, exps


# ── Função principal de scoring ────────────────────────────────────

def score_article(item: dict, cluster_stats: Dict[str, dict],
                  now: datetime) -> Tuple[int, str]:
    """
    Retorna (nota_final, string_de_explicação).
    """
    title = item["title"]
    url = item.get("url", "")
    dt = item.get("dt")
    title_lower = title.lower()

    parts: List[str] = []
    total = 0

    # 1. Tema permanente
    perm_score, perm_exps, max_w = _match_vocab_permanente(title_lower, title)
    perm_score = min(perm_score, 24)
    total += perm_score
    parts.extend(perm_exps)

    # 2. Ações / acontecimento
    acao_bonus, acao_exp = _match_acoes(title_lower, total)
    acao_bonus = min(acao_bonus, 10)
    total += acao_bonus
    if acao_exp:
        parts.append(acao_exp)

    # 3. Impacto
    imp_bonus, imp_exps = _match_impacto(title_lower, total)
    total += imp_bonus
    parts.extend(imp_exps)

    # 4. Especificidade
    esp_bonus, esp_exps = _match_especificidade(title, total)
    total += esp_bonus
    parts.extend(esp_exps)

    # 5. Tendências
    tend_bonus, tend_exps = _tendencia_bonus(title_lower, title, cluster_stats)
    total += tend_bonus
    parts.extend(tend_exps)

    # 6. Editoria
    ed_bonus, ed_exp = _calc_editoria(url)
    total += ed_bonus
    if ed_exp:
        parts.append(ed_exp)

    # 8. Penalidades
    # subtotal_positivo = tudo acima (sem penalidades); has_cluster = tendência detectada
    pen_total, pen_exps = _match_penalidades(
        title_lower,
        subtotal_positivo=total,
        has_cluster=tend_bonus > 0,
    )
    total += pen_total
    parts.extend(pen_exps)

    explanation = " | ".join(parts) if parts else "sem correspondências"
    return total, explanation


def apply_repetition_penalty(top_items: List[dict]) -> List[dict]:
    """
    Penaliza itens muito similares dentro do top 20 de um veículo.
    Critério: ≥4 palavras significativas em comum.
    A penalidade afeta _score mas o _score_selecao (usado para ranking) não muda.
    """
    stop = {"de","do","da","os","as","um","uma","que","em","no","na","e","o","a","é"}
    for i, item in enumerate(top_items):
        words_i = {w for w in item["title"].lower().split() if w not in stop and len(w) > 3}
        for j in range(i):
            words_j = {w for w in top_items[j]["title"].lower().split() if w not in stop and len(w) > 3}
            if len(words_i & words_j) >= 4:
                item["_score"] = item.get("_score", 0) - 5
                item["_explanation"] = item.get("_explanation", "") + " | repetição -5"
    return top_items


def select_top(items: List[dict], cluster_stats: Dict[str, dict],
               now: datetime, n: int = 20) -> List[dict]:
    for item in items:
        s, exp = score_article(item, cluster_stats, now)
        item["_score"] = s
        item["_explanation"] = exp
    items.sort(key=lambda x: x["_score"], reverse=True)
    top = items[:n]
    for rank, item in enumerate(top, 1):
        item["_rank"] = rank
        # Guarda o score usado para seleção (pré-penalidade de repetição)
        # para que a planilha mostre valores comparáveis entre selecionados e não-selecionados
        item["_score_selecao"] = item["_score"]
    apply_repetition_penalty(top)
    return top


# ═══════════════════════════════════════════════════════════════════
# FILTRO TEMPORAL
# ═══════════════════════════════════════════════════════════════════

def get_time_filter() -> Tuple[datetime, Optional[datetime]]:
    print("\n" + "═" * 40)
    print("  CONFIGURAÇÃO DE TEMPO")
    print("═" * 40)
    print("\n  1 - Últimas X horas\n  2 - Janela específica\n")
    while True:
        modo = input("Escolha (1 ou 2): ").strip()
        if modo in ("1", "2"):
            break
        print("  Digite 1 ou 2.")

    if modo == "1":
        while True:
            try:
                horas = int(input("\nQuantas horas para trás? Exemplo: 9\n→ ").strip())
                if horas > 0:
                    break
                print("  Número maior que zero.")
            except ValueError:
                print("  Digite um número inteiro.")
        ws = datetime.now(tz=timezone.utc) - timedelta(hours=horas)
        print(f"\n  Últimas {horas}h (desde {ws.astimezone(LOCAL_TZ).strftime('%d/%m %H:%M')} Brasília)")
        return ws, None
    else:
        fmt = "%d/%m/%Y %H:%M"
        print("\n  Formato: DD/MM/YYYY HH:MM")
        while True:
            try:
                dt_start = datetime.strptime(input("\nInício:\n→ ").strip(), fmt).replace(tzinfo=LOCAL_TZ)
                break
            except ValueError:
                print("  Formato inválido.")
        while True:
            try:
                dt_end = datetime.strptime(input("\nFim:\n→ ").strip(), fmt).replace(tzinfo=LOCAL_TZ)
                if dt_end > dt_start:
                    break
                print("  Data final deve ser posterior.")
            except ValueError:
                print("  Formato inválido.")
        print(f"\n  Janela: {dt_start.strftime('%d/%m/%Y %H:%M')} até {dt_end.strftime('%d/%m/%Y %H:%M')}")
        return dt_start.astimezone(timezone.utc), dt_end.astimezone(timezone.utc)


def in_window(dt: Optional[datetime], ws: datetime, we: Optional[datetime]) -> bool:
    if dt is None:
        return False
    return dt >= ws if we is None else ws <= dt <= we


# ═══════════════════════════════════════════════════════════════════
# DATETIME HELPERS
# ═══════════════════════════════════════════════════════════════════

def now_utc() -> datetime:
    return datetime.now(tz=timezone.utc)


_DATE_MODE: str = "fast"   # definido em main() pela escolha do usuário


def parse_iso(raw: str) -> Optional[datetime]:
    raw = raw.strip()
    if _DATE_MODE == "fast":
        # Modo rápido: strptime simples, ignora offset de timezone
        raw2 = raw.rstrip("Z")
        for fmt in ("%Y-%m-%dT%H:%M:%S.%f", "%Y-%m-%dT%H:%M:%S"):
            try:
                return datetime.strptime(raw2[:26], fmt).replace(tzinfo=timezone.utc)
            except ValueError:
                continue
        return None
    else:
        # Modo preciso: fromisoformat respeita -03:00, +00:00, Z
        if raw.endswith("Z"):
            raw = raw[:-1] + "+00:00"
        try:
            dt = datetime.fromisoformat(raw)
            if dt.tzinfo is None:
                dt = dt.replace(tzinfo=timezone.utc)
            return dt.astimezone(timezone.utc)
        except ValueError:
            pass
        return None


def parse_rss_date(raw: str) -> Optional[datetime]:
    for fmt in ("%a, %d %b %Y %H:%M:%S %z", "%d %b %Y %H:%M:%S %z",
                "%a, %d %b %Y %H:%M:%S %Z", "%d %b %Y %H:%M:%S %Z"):
        try:
            return datetime.strptime(raw.strip(), fmt).astimezone(timezone.utc)
        except ValueError:
            continue
    return None


def parse_relative_time(text: str) -> Optional[datetime]:
    now = now_utc()
    m = re.search(r"há\s+(\d+)\s+(minuto|hora|dia)", text.lower())
    if m:
        n, unit = int(m.group(1)), m.group(2)
        if unit.startswith("minuto"):
            return now - timedelta(minutes=n)
        if unit.startswith("hora"):
            return now - timedelta(hours=n)
        return now - timedelta(days=n)
    if "agora" in text.lower():
        return now
    return None


def extract_date_from_url(url: str) -> Optional[datetime]:
    m = re.search(r"/(\d{4})/(\d{2})/(\d{2})/", url)
    if m:
        try:
            return datetime(int(m.group(1)), int(m.group(2)), int(m.group(3)), tzinfo=timezone.utc)
        except ValueError:
            pass
    return None


def format_time(dt: Optional[datetime]) -> str:
    if dt is None:
        return ""
    return dt.astimezone(LOCAL_TZ).strftime("%H:%M")


# ═══════════════════════════════════════════════════════════════════
# FETCHERS (inalterados)
# ═══════════════════════════════════════════════════════════════════

def fetch_sitemap(url: str) -> list:
    """
    Faz parse de Google News Sitemap XML.
    Formato esperado:
      <url>
        <loc>https://...</loc>
        <lastmod>2026-06-22T09:30:00-03:00</lastmod>
        <news:news>
          <news:title>Título</news:title>
          <news:publication_date>2026-06-22T09:30:00-03:00</news:publication_date>
        </news:news>
      </url>
    """
    import xml.etree.ElementTree as ET
    items = []
    try:
        r = requests.get(url, headers=HEADERS, timeout=20)
        r.raise_for_status()
        root = ET.fromstring(r.content)

        SM   = "http://www.sitemaps.org/schemas/sitemap/0.9"
        NEWS = "http://www.google.com/schemas/sitemap-news/0.9"

        def t(el, tag, ns):
            found = el.find(f"{{{ns}}}{tag}")
            return found.text.strip() if found is not None and found.text else ""

        # Extrai data do sitemap diário datado (ex: /sitemap/2026-06-23/) para
        # usar como fallback quando a URL do artigo não contém data (Estadão).
        _sitemap_date_fallback: Optional[datetime] = None
        _m_sd = re.search(r'/(\d{4}-\d{2}-\d{2})/', url)
        if _m_sd:
            try:
                _BRT = timezone(timedelta(hours=-3))
                _d = datetime.strptime(_m_sd.group(1), '%Y-%m-%d')
                _sitemap_date_fallback = _d.replace(hour=12, minute=0, tzinfo=_BRT).astimezone(timezone.utc)
            except ValueError:
                pass

        for url_el in root.findall(f"{{{SM}}}url"):
            loc = t(url_el, "loc", SM)
            if not loc:
                continue

            title  = ""
            pub_dt = None
            has_pub_date = False

            news_el = url_el.find(f"{{{NEWS}}}news")
            if news_el is not None:
                title    = t(news_el, "title",            NEWS)
                date_raw = t(news_el, "publication_date", NEWS)
                if date_raw:
                    pub_dt = parse_iso(date_raw)
                    has_pub_date = pub_dt is not None

            _date_from_lastmod = False
            if _DATE_MODE == "fast":
                # Modo rápido: valida data da URL, usa lastmod como fallback
                url_date = extract_date_from_url(loc)
                if url_date is not None:
                    if pub_dt is None or abs((pub_dt.date() - url_date.date()).days) > 1:
                        BRT = timezone(timedelta(hours=-3))
                        same = pub_dt is not None and abs((pub_dt.date() - url_date.date()).days) <= 1
                        hour   = pub_dt.astimezone(BRT).hour   if same else 12
                        minute = pub_dt.astimezone(BRT).minute if same else 0
                        pub_dt = url_date.replace(hour=hour, minute=minute, tzinfo=BRT).astimezone(timezone.utc)
                if pub_dt is None:
                    lastmod = t(url_el, "lastmod", SM)
                    if lastmod:
                        pub_dt = parse_iso(lastmod)
                        _date_from_lastmod = True
            else:
                # Modo preciso: lastmod mesmo dia → usa hora; outro dia → meio-dia BRT
                # Artigos sem pub_date serão enriquecidos via página (_needs_enrich)
                if not has_pub_date:
                    url_date    = extract_date_from_url(loc)
                    lastmod_raw = t(url_el, "lastmod", SM)
                    lastmod_dt  = parse_iso(lastmod_raw) if lastmod_raw else None
                    if url_date is not None:
                        BRT = timezone(timedelta(hours=-3))
                        if lastmod_dt is not None and lastmod_dt.date() == url_date.date():
                            pub_dt = lastmod_dt
                        else:
                            pub_dt = url_date.replace(hour=12, minute=0, tzinfo=BRT).astimezone(timezone.utc)
                    else:
                        # Sem data na URL: usa data do sitemap diário (noon BRT) se
                        # disponível, pois lastmod pode refletir edição posterior.
                        if _sitemap_date_fallback is not None:
                            pub_dt = _sitemap_date_fallback
                        else:
                            pub_dt = lastmod_dt

            _needs_title_enrich = False
            if not title:
                slug = loc.rstrip("/").split("/")[-1]
                title = re.sub(r"[-_]", " ", slug).title()
                _needs_title_enrich = True

            if len(title) < 15:
                continue

            items.append({
                "title":               title,
                "url":                 loc.split("?")[0],
                "dt":                  pub_dt,
                "_needs_enrich":       (_DATE_MODE == "accurate" and not has_pub_date),
                "_needs_title_enrich": _needs_title_enrich,
                "_date_from_lastmod":  _date_from_lastmod,
            })

    except Exception as e:
        print(f"  [aviso] Sitemap ({url}): {e}", file=sys.stderr)
    return items


def fetch_rss(url: str) -> list:
    items = []
    try:
        feed = feedparser.parse(url)
        for entry in feed.entries:
            title = entry.get("title", "").strip()
            raw_link = entry.get("link", "").strip()
            # Folha usa redirect: redir.folha.com.br/.../*https://real-url
            if "*https://" in raw_link:
                raw_link = raw_link.split("*https://", 1)[1]
                raw_link = "https://" + raw_link
            link = raw_link.split("?")[0]
            if not title or not link:
                continue
            dt = None
            if hasattr(entry, "published_parsed") and entry.published_parsed:
                dt = datetime.fromtimestamp(calendar.timegm(entry.published_parsed), tz=timezone.utc)
            else:
                raw = entry.get("published", "") or entry.get("updated", "")
                dt = parse_rss_date(raw)
            items.append({"title": title, "url": link, "dt": dt})
    except Exception as e:
        print(f"  [aviso] RSS ({url}): {e}", file=sys.stderr)
    return items


def _fetch_article_pub_date(url: str) -> Optional[datetime]:
    """Busca data de publicação na página do artigo (modo preciso).
    Ordem: JSON-LD datePublished → published_time JSON inline → itemprop datePublished."""
    try:
        r = requests.get(url, headers=HEADERS, timeout=10)
        if r.status_code != 200:
            return None
        html = r.text

        # 1. JSON-LD datePublished (Estadão e outros)
        for raw in re.findall(
            r'<script[^>]+application/ld\+json[^>]*>(.*?)</script>',
            html, re.DOTALL | re.IGNORECASE
        ):
            try:
                data = json.loads(raw.strip())
                entries = data if isinstance(data, list) else [data]
                for obj in entries:
                    if not isinstance(obj, dict):
                        continue
                    if not any(x in str(obj.get("@type", ""))
                               for x in ("Article", "News", "Report")):
                        continue
                    date_raw = obj.get("datePublished")
                    if date_raw:
                        dt = parse_iso(date_raw)
                        if dt:
                            return dt
            except Exception:
                pass

        # 2. published_time em JSON inline (Valor, Globo)
        m = re.search(r'["\']published_time["\']\s*:\s*["\']([^"\']+)["\']', html)
        if m:
            dt = parse_iso(m.group(1))
            if dt:
                return dt

        # 3. itemprop="datePublished" em <time> ou <meta> (fallback Valor/Globo)
        for pattern in [
            r'itemprop=["\']datePublished["\'][^>]+datetime=["\']([^"\']+)["\']',
            r'datetime=["\']([^"\']+)["\'][^>]+itemprop=["\']datePublished["\']',
            r'<meta[^>]+itemprop=["\']datePublished["\'][^>]+content=["\']([^"\']+)["\']',
        ]:
            m = re.search(pattern, html, re.IGNORECASE)
            if m:
                dt = parse_iso(m.group(1))
                if dt:
                    return dt

    except Exception:
        pass
    return None


def _fetch_article_title(url: str) -> Optional[str]:
    """Busca o título real na página do artigo via JSON-LD ou <title>."""
    try:
        r = requests.get(url, headers=HEADERS, timeout=10)
        if r.status_code != 200:
            return None
        html = r.text
        for raw in re.findall(
            r'<script[^>]+application/ld\+json[^>]*>(.*?)</script>',
            html, re.DOTALL | re.IGNORECASE
        ):
            try:
                data = json.loads(raw.strip())
                entries = data if isinstance(data, list) else [data]
                for obj in entries:
                    if not isinstance(obj, dict):
                        continue
                    if not any(x in str(obj.get("@type", ""))
                               for x in ("Article", "News", "Report")):
                        continue
                    headline = obj.get("headline", "").strip()
                    if len(headline) >= 15:
                        return headline
            except Exception:
                pass
        m = re.search(r'<title[^>]*>([^<]{15,})</title>', html, re.IGNORECASE)
        if m:
            return m.group(1).strip().split(" - ")[0].split(" | ")[0].strip()
    except Exception:
        pass
    return None


def _enrich_article_titles(items: List[dict]) -> None:
    """Busca títulos reais para artigos marcados com _needs_title_enrich=True."""
    from concurrent.futures import ThreadPoolExecutor, as_completed
    to_enrich = [item for item in items if item.get("_needs_title_enrich")]
    if not to_enrich:
        return
    print(f"    Buscando títulos em {len(to_enrich)} páginas...", end=" ", flush=True)
    with ThreadPoolExecutor(max_workers=20) as executor:
        futures = {executor.submit(_fetch_article_title, item["url"]): item
                   for item in to_enrich}
        for future in as_completed(futures):
            item = futures[future]
            title = future.result()
            if title:
                item["title"] = title
            item.pop("_needs_title_enrich", None)
    print("concluído.")


def _enrich_article_dates(items: List[dict]) -> None:
    """
    Busca datePublished nas páginas de artigos marcados com _needs_enrich=True.
    Usa até 20 threads em paralelo para minimizar tempo de espera.
    """
    from concurrent.futures import ThreadPoolExecutor, as_completed
    to_enrich = [item for item in items if item.get("_needs_enrich")]
    if not to_enrich:
        return
    print(f"    Buscando datas em {len(to_enrich)} páginas...", end=" ", flush=True)
    with ThreadPoolExecutor(max_workers=20) as executor:
        futures = {executor.submit(_fetch_article_pub_date, item["url"]): item
                   for item in to_enrich}
        for future in as_completed(futures):
            item = futures[future]
            dt = future.result()
            if dt is not None:
                item["dt"] = dt
            item.pop("_needs_enrich", None)
    print("concluído.")


# ═══════════════════════════════════════════════════════════════════
# MOTOR PLAYWRIGHT (compartilhado entre todos os fetchers HTML)
# ═══════════════════════════════════════════════════════════════════
# Instalar uma vez no terminal:
#   pip3 install playwright
#   playwright install chromium

_pw = None
_pw_browser = None
_pw_ctx = None


def pw_start():
    global _pw, _pw_browser, _pw_ctx
    from playwright.sync_api import sync_playwright
    _pw = sync_playwright().start()
    _pw_browser = _pw.chromium.launch(
        headless=True,
        args=[
            "--disable-blink-features=AutomationControlled",
            "--no-sandbox",
            "--disable-dev-shm-usage",
        ],
    )
    _pw_ctx = _pw_browser.new_context(
        user_agent=(
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/124.0.0.0 Safari/537.36"
        ),
        locale="pt-BR",
        timezone_id="America/Sao_Paulo",
        viewport={"width": 1280, "height": 900},
        extra_http_headers={
            "Accept-Language": "pt-BR,pt;q=0.9,en;q=0.8",
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        },
    )
    # Remove o flag webdriver que os sites detectam
    _pw_ctx.add_init_script(
        "Object.defineProperty(navigator, 'webdriver', {get: () => undefined})"
    )
    print("  [playwright] Navegador iniciado.")


def pw_stop():
    global _pw, _pw_browser, _pw_ctx
    try:
        if _pw_browser:
            _pw_browser.close()
        if _pw:
            _pw.stop()
    except Exception:
        pass
    _pw = _pw_browser = _pw_ctx = None


def pw_html(url: str, wait_selector: str = None) -> str:
    """Abre url no Chromium headless e retorna o HTML totalmente renderizado."""
    if _pw_ctx is None:
        raise RuntimeError("Playwright não iniciado — chame pw_start() primeiro.")
    page = _pw_ctx.new_page()
    try:
        # Bloqueia recursos pesados que não precisamos (rastreadores, mídia)
        page.route(
            re.compile(r"\.(png|jpg|jpeg|gif|svg|webp|woff2?|mp4|mp3)(\?.*)?$"),
            lambda route: route.abort(),
        )
        page.goto(url, wait_until="domcontentloaded", timeout=30000)
        # Espera artigos aparecerem ou aguarda tempo fixo
        if wait_selector:
            try:
                page.wait_for_selector(wait_selector, timeout=10000)
            except Exception:
                page.wait_for_timeout(3000)
        else:
            page.wait_for_timeout(3000)
        # Rola a página para forçar lazy-load de conteúdo
        page.evaluate("window.scrollTo(0, document.body.scrollHeight / 2)")
        page.wait_for_timeout(1000)
        return page.content()
    except Exception as e:
        print(f"  [aviso playwright] {url}: {e}", file=sys.stderr)
        return ""
    finally:
        page.close()


def _extract_json_ld(html: str) -> List[dict]:
    """
    Extrai artigos de blocos JSON-LD <script type='application/ld+json'>.
    Retorna lista de {title, url, dt} — método mais confiável quando disponível.
    """
    items = []
    seen: set = set()
    pattern = r'<script[^>]+type=["\']application/ld\+json["\'][^>]*>(.*?)</script>'
    for raw in re.findall(pattern, html, re.DOTALL | re.IGNORECASE):
        try:
            data = json.loads(raw.strip())
        except Exception:
            continue
        entries = data if isinstance(data, list) else [data]
        for obj in entries:
            if not isinstance(obj, dict):
                continue
            if obj.get("@type") not in ("NewsArticle", "Article", "ReportageNewsArticle", "WebPage"):
                continue
            title = obj.get("headline", "")
            url = obj.get("url", "")
            if not url:
                mep = obj.get("mainEntityOfPage")
                url = mep.get("@id", "") if isinstance(mep, dict) else (mep or "")
            date_raw = obj.get("datePublished") or obj.get("dateModified") or ""
            if title and url and len(title) >= 15 and url not in seen:
                seen.add(url)
                items.append({"title": title, "url": url.split("?")[0],
                               "dt": parse_iso(date_raw) if date_raw else None})
    return items


def _extract_time_tags(html: str) -> Dict[str, datetime]:
    """
    Mapeia URL → datetime extraindo <time datetime="..."> do HTML.
    Usado para enriquecer datas quando o parser principal não as encontra.
    """
    soup = BeautifulSoup(html, "html.parser")
    url_to_dt: Dict[str, datetime] = {}
    for time_tag in soup.find_all("time", attrs={"datetime": True}):
        dt = parse_iso(time_tag["datetime"])
        if dt is None:
            continue
        # Sobe na árvore para achar o link mais próximo
        for parent in time_tag.parents:
            a = parent.find("a", href=True) if hasattr(parent, "find") else None
            if a and a["href"].startswith("http"):
                url_to_dt[a["href"].split("?")[0]] = dt
                break
    return url_to_dt


# ── Fetchers HTML (agora via Playwright) ──────────────────────────

def fetch_estadao(url: str) -> list:
    html = pw_html(url, wait_selector="article, h2, h3")
    if not html:
        return []

    # 1. JSON-LD (mais confiável)
    items = _extract_json_ld(html)

    # 2. Fallback: regex no JSON embutido (Arc Publishing)
    if not items:
        pattern = (r'"_id":"[^"]+"[^{]*?"type":"story".*?'
                   r'"first_publish_date":"(\d{4}-\d{2}-\d{2}T[^"]+)".*?'
                   r'"headlines":\{"basic":"([^"]+)".*?'
                   r'"canonical_url":"(/[^"]+)"')
        seen: set = set()
        for date_raw, title, path in re.findall(pattern, html, re.DOTALL):
            fu = "https://www.estadao.com.br" + path
            if fu not in seen and len(title) >= 15:
                seen.add(fu)
                items.append({"title": title, "url": fu, "dt": parse_iso(date_raw)})

    # 3. Enriquece datas faltantes com <time datetime>
    if items:
        time_map = _extract_time_tags(html)
        for item in items:
            if item["dt"] is None:
                item["dt"] = time_map.get(item["url"]) or extract_date_from_url(item["url"])

    if not items:
        print(f"  [debug] Estadão: 0 artigos em {url} "
              f"(HTML {len(html)} chars)", file=sys.stderr)
    return items


def fetch_oglobo(url: str) -> list:
    html = pw_html(url, wait_selector="article, h2, h3")
    if not html:
        return []

    # 1. JSON-LD
    items = _extract_json_ld(html)

    # 2. Fallback: regex no JSON embutido
    if not items:
        pattern = (r'"title":"([^"]{15,}?)".*?'
                   r'"url":"(https://oglobo\.globo\.com/[^"]+\.ghtml[^"]*)".*?'
                   r'"created":"(\d{4}-\d{2}-\d{2}T\d{2}:\d{2}:\d{2})')
        seen: set = set()
        for title, art_url, date_raw in re.findall(pattern, html, re.DOTALL):
            art_url = art_url.split("?")[0]
            if art_url not in seen and len(title) >= 15:
                seen.add(art_url)
                items.append({"title": title, "url": art_url, "dt": parse_iso(date_raw)})

    # 3. Enriquece datas faltantes
    if items:
        time_map = _extract_time_tags(html)
        for item in items:
            if item["dt"] is None:
                item["dt"] = time_map.get(item["url"]) or extract_date_from_url(item["url"])

    if not items:
        print(f"  [debug] O Globo: 0 artigos em {url} "
              f"(HTML {len(html)} chars)", file=sys.stderr)
    return items


def fetch_valor(url: str) -> list:
    html = pw_html(url, wait_selector=".feed-post-body, article, h2")
    if not html:
        return []

    items = []
    soup = BeautifulSoup(html, "html.parser")
    seen: set = set()

    # 1. Estrutura feed-post-body (Valor usa server-side rendering com Globo CMS)
    for post in soup.find_all(class_="feed-post-body"):
        a = next((c for c in post.find_all("a", href=True)
                  if len(c.get_text(strip=True)) > 20
                  and "valor.globo.com" in c["href"]
                  and "/noticia/" in c["href"]), None)
        if not a:
            continue
        title = a.get_text(strip=True)
        art_url = a["href"].split("?")[0]
        if art_url in seen:
            continue
        seen.add(art_url)
        dt = None
        # <time datetime="..."> é mais confiável que texto relativo
        time_tag = post.find("time", attrs={"datetime": True})
        if time_tag:
            dt = parse_iso(time_tag["datetime"])
        if dt is None:
            ts = post.find(class_="feed-post-datetime")
            if ts:
                dt = parse_relative_time(ts.get_text(strip=True))
        if dt is None:
            dt = extract_date_from_url(art_url)
        items.append({"title": title, "url": art_url, "dt": dt})

    # 2. Fallback: JSON-LD
    if not items:
        items = _extract_json_ld(html)
        time_map = _extract_time_tags(html)
        for item in items:
            if item["dt"] is None:
                item["dt"] = time_map.get(item["url"]) or extract_date_from_url(item["url"])

    if not items:
        print(f"  [debug] Valor: 0 artigos em {url} "
              f"(HTML {len(html)} chars)", file=sys.stderr)
    return items


FETCHERS = {
    "rss":          fetch_rss,
    "sitemap":      fetch_sitemap,
    "estadao_html": fetch_estadao,
    "globo_html":   fetch_oglobo,
    "valor_html":   fetch_valor,
}


def _daily_sitemap_urls(daily_cfg: dict, window_start: datetime,
                         window_end: Optional[datetime]) -> List[str]:
    """
    Gera URLs de sitemaps diários para todos os dias que se sobrepõem à janela.

    Os arquivos são nomeados por data BRT (UTC-3), independente do veículo.
    As datas nas entradas de cada arquivo têm offset explícito:
      - Globo/Valor: -03:00  →  parse_iso já lida corretamente
      - Estadão (Arc): Z (UTC) →  parse_iso já lida corretamente
    """
    BRT = timezone(timedelta(hours=-3))
    we = window_end or datetime.now(tz=timezone.utc)

    # Dias necessários em BRT
    d = window_start.astimezone(BRT).replace(hour=0, minute=0, second=0, microsecond=0)
    end_d = we.astimezone(BRT).replace(hour=0, minute=0, second=0, microsecond=0)
    dates: List[datetime] = []
    while d <= end_d:
        dates.append(d)
        d += timedelta(days=1)

    urls: List[str] = []
    fmt  = daily_cfg["format"]
    base = daily_cfg["base"]

    for date in dates:
        if fmt == "globo":
            # Formato: base/{ano}/{mes:02d}/{dia:02d}_{pagina}.xml
            # Podem existir várias páginas por dia (_1, _2, …) — buscamos até vazia
            for page in range(1, 10):
                urls.append(
                    f"{base}/{date.year}/{date.month:02d}/{date.day:02d}_{page}.xml"
                )
        elif fmt == "estadao":
            # Formato: base/{ano}-{mes:02d}-{dia:02d}/?outputType=xml  (1 arquivo/dia)
            urls.append(f"{base}/{date.strftime('%Y-%m-%d')}/?outputType=xml")

    return urls


def fetch_source(name: str, config: dict,
                 window_start: Optional[datetime] = None,
                 window_end: Optional[datetime] = None) -> list:
    all_items: List[dict] = []
    seen_urls: set = set()

    def _add(items: list):
        for item in items:
            if item["url"] not in seen_urls:
                seen_urls.add(item["url"])
                item["source"] = name
                all_items.append(item)

    src_type = config["type"]

    if src_type == "sitemap_daily":
        # 1. URLs estáticas (news.xml, RSS)
        for url in config.get("urls", []):
            print(f"  Buscando: {url}")
            try:
                fn = fetch_rss if ("rss" in url or "feeds.folha" in url) else fetch_sitemap
                items = fn(url)
            except Exception as e:
                print(f"  [erro] {url}: {e}", file=sys.stderr)
                items = []
            print(f"    → {len(items)} itens")
            _add(items)

        # 2. Sitemaps diários baseados na janela
        if "daily" in config and window_start is not None:
            day_urls = _daily_sitemap_urls(config["daily"], window_start, window_end)
            fmt = config["daily"]["format"]
            i = 0
            while i < len(day_urls):
                url = day_urls[i]
                print(f"  Buscando: {url}")
                try:
                    items = fetch_sitemap(url)
                except Exception as e:
                    print(f"  [erro] {url}: {e}", file=sys.stderr)
                    items = []
                print(f"    → {len(items)} itens")

                # Para formato globo, se página N devolver 0, para para esse dia
                if fmt == "globo" and not items:
                    # Pula as restantes páginas do mesmo dia
                    current_prefix = re.sub(r'_\d+\.xml$', '', url)
                    while i + 1 < len(day_urls) and day_urls[i + 1].startswith(current_prefix):
                        i += 1

                _add(items)
                i += 1
    else:
        fn = FETCHERS[src_type]
        for url in config["urls"]:
            print(f"  Buscando: {url}")
            try:
                items = fn(url)
            except Exception as e:
                print(f"  [erro] {url}: {e}", file=sys.stderr)
                items = []
            print(f"    → {len(items)} itens")
            _add(items)

    # Enriquece títulos ausentes no sitemap (ex: Estadão) — sempre, em paralelo
    _enrich_article_titles(all_items)

    # Modo preciso: enriquece datas buscando páginas em paralelo
    if src_type == "sitemap_daily" and _DATE_MODE == "accurate":
        _enrich_article_dates(all_items)

    # Remove flags internos antes de retornar
    for item in all_items:
        item.pop("_needs_enrich", None)
        item.pop("_needs_title_enrich", None)

    return all_items


# ═══════════════════════════════════════════════════════════════════
# PLANILHA
# ═══════════════════════════════════════════════════════════════════

def generate_spreadsheet(audit_rows: list, cluster_stats: Dict[str, dict],
                          filename: str = None, return_workbook: bool = False):
    import os
    if filename is None:
        ts = datetime.now(tz=LOCAL_TZ).strftime("%Y%m%d_%H%M")
        filename = os.path.join(os.path.expanduser("~"), "Desktop", f"manchetes_auditoria_{ts}.xlsx")
    wb = openpyxl.Workbook()

    # ── Sheet 1: Clusters do Dia ───────────────────────────────────
    ws_cl = wb.active
    ws_cl.title = "Clusters do Dia"

    hfont    = Font(bold=True, color="FFFFFF")
    hfill    = PatternFill("solid", fgColor="1F4E79")
    hfill_g  = PatternFill("solid", fgColor="375623")  # verde escuro p/ grupais
    hfill_h  = PatternFill("solid", fgColor="7B3F00")  # marrom p/ histórico puro
    sec_font = Font(bold=True, color="FFFFFF")
    wrap     = Alignment(wrap_text=True, vertical="top")
    center   = Alignment(horizontal="center", vertical="center")

    cl_headers = [
        "Cluster / Tendência", "Tipo", "Membros agrupados",
        "Art. (janela)", "Art. (histórico)",
        "Veículos (janela)", "Veículos (histórico)",
        "Pontos", "Explicação da pontuação",
    ]
    for c, h in enumerate(cl_headers, 1):
        cell = ws_cl.cell(row=1, column=c, value=h)
        cell.font = hfont; cell.fill = hfill; cell.alignment = center

    def _write_section_header(ws, row, label, fill):
        cell = ws.cell(row=row, column=1, value=label)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = fill
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(cl_headers))

    def _cluster_row(ws, row_i, cname, st):
        bonus    = st["bonus"]
        is_g     = st["is_grouped"]
        is_honly = st["count_current"] == 0
        if is_honly:
            bg = "FCE4D6"
        elif is_g:
            bg = "C6EFCE" if bonus >= 6 else "EBF3E8"
        else:
            bg = "D9EAD3" if bonus >= 6 else "EBF3E8" if bonus >= 3 else "EDEDED"
        row_fill = PatternFill("solid", fgColor=bg)

        tipo = "Grupo" if is_g else ("Histórico puro" if is_honly else "Individual")
        members_str = ", ".join(st.get("group_members", [])) if is_g else ""
        src_j = ", ".join(sorted(st["sources_current"]))
        src_h = ", ".join(sorted(st["sources_historical"]))

        vals = [
            cname, tipo, members_str,
            st["count_current"], st["count_historical"],
            src_j, src_h,
            bonus, st["bonus_explanation"],
        ]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=row_i, column=c, value=v)
            cell.fill = row_fill; cell.alignment = wrap

    row_i = 2
    grouped_clusters  = {k: v for k, v in cluster_stats.items() if v["is_grouped"]}
    individ_clusters  = {k: v for k, v in cluster_stats.items() if not v["is_grouped"] and v["count_current"] > 0}
    historic_clusters = {k: v for k, v in cluster_stats.items() if not v["is_grouped"] and v["count_current"] == 0}

    sort_key = lambda x: (-x[1]["bonus"], -x[1]["count"])

    if grouped_clusters:
        _write_section_header(ws_cl, row_i, "▌ CLUSTERS GRUPAIS (co-ocorrência)", hfill_g)
        row_i += 1
        for cname, st in sorted(grouped_clusters.items(), key=sort_key):
            _cluster_row(ws_cl, row_i, cname, st)
            row_i += 1
        row_i += 1  # linha em branco

    _write_section_header(ws_cl, row_i, "▌ CLUSTERS INDIVIDUAIS (janela atual)", hfill)
    row_i += 1
    for cname, st in sorted(individ_clusters.items(), key=sort_key):
        _cluster_row(ws_cl, row_i, cname, st)
        row_i += 1

    if historic_clusters:
        row_i += 1
        _write_section_header(ws_cl, row_i, "▌ CLUSTERS HISTÓRICOS (só fora da janela)", hfill_h)
        row_i += 1
        for cname, st in sorted(historic_clusters.items(), key=sort_key):
            _cluster_row(ws_cl, row_i, cname, st)
            row_i += 1

    col_widths = [28, 16, 45, 12, 14, 30, 30, 8, 70]
    for c, w in enumerate(col_widths, 1):
        ws_cl.column_dimensions[openpyxl.utils.get_column_letter(c)].width = w

    # ── Sheet 2: Manchetes ─────────────────────────────────────────
    ws_art = wb.create_sheet("Manchetes")

    art_headers = [
        "Veículo", "Tópico", "Título", "URL", "Horário", "Nota Seleção", "Rank",
        "Ação", "Impacto", "Especif.", "Tendência", "Editoria", "Penalidades", "Repetição",
        "Status", "Explicação completa",
    ]
    for c, h in enumerate(art_headers, 1):
        cell = ws_art.cell(row=1, column=c, value=h)
        cell.font = hfont; cell.fill = hfill; cell.alignment = center

    status_fills = {
        "Selecionado":               PatternFill("solid", fgColor="C6EFCE"),
        "Top 20 — disponível":       PatternFill("solid", fgColor="EBF3E8"),
        "Fora da janela de tempo":   PatternFill("solid", fgColor="FCE4D6"),
        "Sem data identificável":    PatternFill("solid", fgColor="F2DCDB"),
        "Score baixo / não entrou":  PatternFill("solid", fgColor="EDEDED"),
    }

    for row_idx, row in enumerate(audit_rows, start=2):
        fill = status_fills.get(row.get("status", ""), PatternFill())
        vals = [
            row.get("source", ""),
            row.get("topico", ""),
            row.get("title", ""),
            row.get("url", ""),
            row.get("horario", ""),
            row.get("score", ""),
            row.get("rank", ""),
            row.get("s_acao", ""),
            row.get("s_impacto", ""),
            row.get("s_especif", ""),
            row.get("s_tendencia", ""),
            row.get("s_editoria", ""),
            row.get("s_penalidade", ""),
            row.get("s_repeticao", ""),
            row.get("status", ""),
            row.get("explicacao", ""),
        ]
        for c, v in enumerate(vals, 1):
            cell = ws_art.cell(row=row_idx, column=c, value=v)
            cell.fill = fill; cell.alignment = wrap

    widths = [18, 14, 55, 45, 8, 8, 6, 6, 8, 8, 10, 10, 12, 8, 24, 80]
    for c, w in enumerate(widths, 1):
        ws_art.column_dimensions[openpyxl.utils.get_column_letter(c)].width = w

    ws_art.freeze_panes = "A2"
    ws_art.auto_filter.ref = f"A1:P{len(audit_rows)+1}"

    if return_workbook:
        return wb
    wb.save(filename)
    print(f"📊 Planilha salva em {filename}")


# ═══════════════════════════════════════════════════════════════════
# EXIBIÇÃO
# ═══════════════════════════════════════════════════════════════════

def print_section(name: str, prefix: str, items: list):
    print(f"\n{'─'*55}")
    print(f"  {name}  ({len(items)} manchetes)")
    print(f"{'─'*55}")
    for idx, item in enumerate(items, 1):
        time_str = format_time(item.get("dt"))
        tp = f" [{time_str}]" if time_str else ""
        print(f"({prefix}{idx}) [{item.get('_score',0):+d}] {item['title']}{tp}")
        print(f"     {item['url']}")


# ═══════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════

def _shorten_url(url: str) -> str:
    """Encurta uma URL via TinyURL. Retorna a original em caso de falha."""
    import urllib3
    urllib3.disable_warnings()
    try:
        r = requests.get(
            "https://tinyurl.com/api-create.php",
            params={"url": url},
            timeout=8,
            verify=False,
        )
        if r.status_code == 200 and r.text.startswith("https://"):
            return r.text.strip()
    except Exception:
        pass
    return url


# ═══════════════════════════════════════════════════════════════════
# RETOOL RPC — funções expostas para a interface web
# ═══════════════════════════════════════════════════════════════════

# Estado compartilhado entre chamadas RPC (buscar → gerar clipping/planilha)
_retool_state: dict = {
    "audit_rows":    [],
    "cluster_stats": {},
    "all_selected":  {},
}


def _retool_run_pipeline(modo: str, horas: int = None,
                          janela_inicio: str = None, janela_fim: str = None) -> None:
    """Executa fetch + scoring e salva resultado em _retool_state."""
    global _DATE_MODE
    _DATE_MODE = "fast"

    if modo == "horas":
        window_start = datetime.now(tz=timezone.utc) - timedelta(hours=int(horas))
        window_end   = None
    else:
        fmt = "%d/%m/%Y %H:%M"
        window_start = datetime.strptime(janela_inicio, fmt).replace(tzinfo=LOCAL_TZ).astimezone(timezone.utc)
        window_end   = datetime.strptime(janela_fim,    fmt).replace(tzinfo=LOCAL_TZ).astimezone(timezone.utc)

    now = now_utc()
    all_fetched: List[dict] = []
    fetched_by_source: Dict[str, List[dict]] = {}

    for name, config in SOURCES.items():
        try:
            items = fetch_source(name, config, window_start, window_end)
        except Exception:
            items = []
        fetched_by_source[name] = items
        all_fetched.extend(items)

    all_in_window = [a for a in all_fetched if in_window(a.get("dt"), window_start, window_end)]
    week_ago = window_start - timedelta(days=7)
    all_historical = [
        a for a in all_fetched
        if not in_window(a.get("dt"), window_start, window_end)
        and a.get("dt") is not None
        and a["dt"] >= week_ago
    ]
    print(f"\n  Calculando clusters e tendências...")
    cluster_stats = build_cluster_stats(all_in_window, all_historical)
    n_clusters = sum(1 for v in cluster_stats.values() if v.get("count_current", 0) > 0)
    n_grouped  = sum(1 for v in cluster_stats.values() if v.get("is_grouped"))
    print(f"  {len(all_in_window)} artigos na janela · {len(all_historical)} históricos")
    print(f"  {n_clusters} clusters ativos · {n_grouped} agrupados")

    audit_rows: List[dict] = []
    all_selected: Dict[str, List[dict]] = {}

    for name, config in SOURCES.items():
        items = fetched_by_source.get(name, [])
        in_win, out_of_win, no_date = [], [], []
        for item in items:
            dt = item.get("dt")
            if dt is None:
                no_date.append(item)
            elif in_window(dt, window_start, window_end):
                in_win.append(item)
            else:
                out_of_win.append(item)

        for item in no_date:
            s, exp = score_article(item, cluster_stats, now)
            audit_rows.append({
                "source": name, "title": item["title"], "url": item["url"],
                "horario": "", "score": s, "topico": _article_topic(item["url"]) or "",
                "s_perm": "", "s_acao": "", "s_impacto": "", "s_especif": "",
                "s_tendencia": "", "s_recencia": "", "s_editoria": "", "s_penalidade": "",
                "status": "Sem data identificável",
                "explicacao": "Sem data — excluído | " + exp, "_item_ref": item,
            })
        for item in out_of_win:
            s, exp = score_article(item, cluster_stats, now)
            dt_str = item["dt"].astimezone(LOCAL_TZ).strftime("%d/%m/%Y %H:%M") if item.get("dt") else ""
            audit_rows.append({
                "source": name, "title": item["title"], "url": item["url"],
                "horario": format_time(item.get("dt")), "score": s,
                "topico": _article_topic(item["url"]) or "",
                "s_perm": "", "s_acao": "", "s_impacto": "", "s_especif": "",
                "s_tendencia": "", "s_recencia": "", "s_editoria": "", "s_penalidade": "",
                "status": "Fora da janela de tempo",
                "explicacao": f"Publicado {dt_str} — fora da janela | {exp}", "_item_ref": item,
            })

        top = select_top(in_win, cluster_stats, now, 30)
        all_selected[name] = top

        for item in in_win:
            s_selecao = item.get("_score_selecao", item.get("_score", 0))
            s_final   = item.get("_score", 0)
            exp = item.get("_explanation", "")
            rank = item.get("_rank")
            def _extract(prefix_str, _exp=exp):
                parts = [p for p in _exp.split(" | ") if prefix_str in p]
                return "; ".join(parts)
            rep_penalty = s_final - s_selecao
            audit_rows.append({
                "source": name, "title": item["title"], "url": item["url"],
                "horario": format_time(item.get("dt")),
                "topico": _article_topic(item["url"]) or "",
                "score": s_selecao, "rank": rank if rank else "",
                "s_acao": _extract("ação"), "s_impacto": _extract("impacto"),
                "s_especif": _extract("especif") or _extract("percentual") or _extract("numérico") or _extract("manchete obj") or _extract("prazo"),
                "s_tendencia": _extract("tendência"), "s_editoria": _extract("editoria"),
                "s_penalidade": _extract("-"),
                "s_repeticao": str(rep_penalty) if rep_penalty < 0 else "",
                "status": "Top 20 — disponível" if rank else "Score baixo / não entrou",
                "explicacao": exp, "_item_ref": item,
            })

    _retool_state["audit_rows"]    = audit_rows
    _retool_state["cluster_stats"] = cluster_stats
    _retool_state["all_selected"]  = all_selected


def retool_buscar_manchetes(modo: str, horas: int = None,
                             janela_inicio: str = None, janela_fim: str = None) -> dict:
    """
    Roda o pipeline e retorna top 30 por veículo.
    modo: "horas" ou "janela"
    horas: int (se modo=="horas")
    janela_inicio / janela_fim: "DD/MM/YYYY HH:MM" (se modo=="janela")
    Retorna: {"Valor Econômico": [{"rank":1,"title":"...","url":"...","score":10,"horario":"14:30"}, ...], ...}
    """
    _retool_run_pipeline(modo, horas, janela_inicio, janela_fim)
    result = {}
    for name, items in _retool_state["all_selected"].items():
        result[name] = [
            {
                "rank":              item.get("_rank", idx + 1),
                "title":             item["title"],
                "url":               item["url"],
                "score":             item.get("_score", 0),
                "horario":           format_time(item.get("dt")),
                "date_from_lastmod": item.get("_date_from_lastmod", False),
            }
            for idx, item in enumerate(items)
        ]
    return result


def retool_gerar_clipping(selecoes: dict) -> dict:
    """
    Gera o clipping final com as manchetes selecionadas.
    selecoes: {"Valor Econômico": ["url1", "url2"], "Estadão": [...], ...}
    Retorna: {"clipping": "texto formatado"}
    """
    all_selected = _retool_state["all_selected"]
    audit_rows   = _retool_state["audit_rows"]

    grouped: Dict[str, List[dict]] = {name: [] for name in SOURCES}
    for name, urls in selecoes.items():
        url_set = set(urls)
        for item in all_selected.get(name, []):
            if item["url"] in url_set:
                grouped[name].append(item)

    selected_urls = {url for urls in selecoes.values() for url in urls}
    for row in audit_rows:
        if row.get("_item_ref", {}).get("url") in selected_urls:
            row["status"] = "Selecionado"

    today_str = datetime.now(tz=LOCAL_TZ).strftime("%d/%m/%Y")
    lines = [
        "*Necton Markets | Clipping de notícias*", "",
        f"*{today_str}*", "",
        "_*Confira as principais notícias dos jornais mais relevantes do país*_", "",
    ]
    for name in SOURCES:
        items = grouped.get(name, [])
        if not items:
            continue
        lines.append(f"*{name}*")
        lines.append("")
        for item in items:
            short_url = _shorten_url(item["url"])
            lines.append(f"- {item['title']} {short_url}")
        lines.append("")
    lines.append("*Necton Investimentos*")
    return {"clipping": "\n".join(lines)}


def retool_gerar_planilha(selecoes: dict) -> dict:
    """
    Gera a planilha de auditoria e retorna em base64 para download no Retool.
    Retorna: {"xlsx_base64": "...", "filename": "manchetes_auditoria_YYYYMMDD_HHMM.xlsx"}
    """
    import base64, io as _io

    audit_rows    = _retool_state["audit_rows"]
    cluster_stats = _retool_state["cluster_stats"]

    selected_urls = {url for urls in selecoes.values() for url in urls}
    for row in audit_rows:
        if row.get("_item_ref", {}).get("url") in selected_urls:
            row["status"] = "Selecionado"
    for row in audit_rows:
        row.pop("_item_ref", None)

    status_order = {"Selecionado": 0, "Top 20 — disponível": 1,
                    "Score baixo / não entrou": 2, "Fora da janela de tempo": 3,
                    "Sem data identificável": 4}
    audit_rows.sort(key=lambda r: (
        list(SOURCES.keys()).index(r["source"]) if r["source"] in SOURCES else 99,
        status_order.get(r.get("status", ""), 5),
        -int(r["score"]) if isinstance(r.get("score"), int) else 0,
    ))

    wb = generate_spreadsheet(audit_rows, cluster_stats, return_workbook=True)
    buf = _io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    ts = datetime.now(tz=LOCAL_TZ).strftime("%Y%m%d_%H%M")
    return {
        "xlsx_base64": base64.b64encode(buf.read()).decode(),
        "filename":    f"manchetes_auditoria_{ts}.xlsx",
    }


def main():
    global _DATE_MODE
    print("\n" + "═" * 40)
    print("  MODO DE ANÁLISE DE DATA")
    print("═" * 40)
    print("\n  1 - Rápido")
    print("  2 - Preciso (busca data de publicação nas páginas — mais lento)\n")
    while True:
        modo_data = input("Escolha (1 ou 2): ").strip()
        if modo_data in ("1", "2"):
            break
        print("  Digite 1 ou 2.")
    _DATE_MODE = "fast" if modo_data == "1" else "accurate"

    window_start, window_end = get_time_filter()
    now = now_utc()

    print("\n🔍 Buscando manchetes...\n")
    print("  Artigos sem data identificável são excluídos.\n")

    # ── Inicia Playwright apenas se houver fonte HTML ──────────────
    html_types = {"estadao_html", "globo_html", "valor_html"}
    needs_browser = any(c["type"] in html_types for c in SOURCES.values())
    if needs_browser:
        try:
            pw_start()
        except Exception as e:
            print(f"\n[ERRO] Playwright não disponível: {e}", file=sys.stderr)
            print("  Instale com:\n    pip3 install playwright\n    py -m playwright install chromium", file=sys.stderr)
            sys.exit(1)

    # ── Fase 1: fetch de tudo ──────────────────────────────────────
    all_fetched: List[dict] = []
    fetched_by_source: Dict[str, List[dict]] = {}

    try:
        for name, config in SOURCES.items():
            print(f"[{name}]")
            try:
                items = fetch_source(name, config, window_start, window_end)
            except Exception as e:
                print(f"  [erro] {name}: {e}", file=sys.stderr)
                items = []
            fetched_by_source[name] = items
            all_fetched.extend(items)
    finally:
        if needs_browser:
            pw_stop()

    # ── Fase 2: clusters ───────────────────────────────────────────
    print("\n  Calculando clusters e tendências...")
    all_in_window = [a for a in all_fetched if in_window(a.get("dt"), window_start, window_end)]
    week_ago = window_start - timedelta(days=7)
    all_historical = [
        a for a in all_fetched
        if not in_window(a.get("dt"), window_start, window_end)
        and a.get("dt") is not None
        and a["dt"] >= week_ago
    ]
    cluster_stats = build_cluster_stats(all_in_window, all_historical)
    active = {k: v for k, v in cluster_stats.items() if v["count"] >= 2}
    if active:
        print(f"  Clusters ativos: {', '.join(active.keys())}")

    # ── Fase 3: scoring, filtro, seleção por fonte ─────────────────
    audit_rows: List[dict] = []
    all_selected: Dict[str, List[dict]] = {}
    code_map: Dict[str, dict] = {}

    for name, config in SOURCES.items():
        items = fetched_by_source.get(name, [])

        in_win, out_of_win, no_date = [], [], []
        for item in items:
            dt = item.get("dt")
            if dt is None:
                no_date.append(item)
            elif in_window(dt, window_start, window_end):
                in_win.append(item)
            else:
                out_of_win.append(item)

        # Registra excluídos por data
        for item in no_date:
            s, exp = score_article(item, cluster_stats, now)
            audit_rows.append({
                "source": name, "title": item["title"], "url": item["url"],
                "horario": "", "score": s,
                "topico": _article_topic(item["url"]) or "",
                "s_perm": "", "s_acao": "", "s_impacto": "", "s_especif": "",
                "s_tendencia": "", "s_recencia": "", "s_editoria": "", "s_penalidade": "",
                "status": "Sem data identificável",
                "explicacao": "Sem data — excluído | " + exp,
                "_item_ref": item,
            })
        for item in out_of_win:
            s, exp = score_article(item, cluster_stats, now)
            dt_str = item["dt"].astimezone(LOCAL_TZ).strftime("%d/%m/%Y %H:%M") if item.get("dt") else ""
            audit_rows.append({
                "source": name, "title": item["title"], "url": item["url"],
                "horario": format_time(item.get("dt")), "score": s,
                "topico": _article_topic(item["url"]) or "",
                "s_perm": "", "s_acao": "", "s_impacto": "", "s_especif": "",
                "s_tendencia": "", "s_recencia": "", "s_editoria": "", "s_penalidade": "",
                "status": "Fora da janela de tempo",
                "explicacao": f"Publicado {dt_str} — fora da janela | {exp}",
                "_item_ref": item,
            })

        # Seleciona top 20
        top = select_top(in_win, cluster_stats, now, 20)
        top_url_rank = {i["url"]: i["_rank"] for i in top}
        all_selected[name] = top
        prefix = config["prefix"]

        for idx, item in enumerate(top, 1):
            code_map[f"{prefix}{idx}"] = {"source": name, **item}

        # Registra itens dentro da janela com detalhes de score
        for item in in_win:
            # _score_selecao = score antes da penalidade de repetição (base de comparação justa)
            # _score = score final (pode ser menor por penalidade pós-seleção)
            s_selecao = item.get("_score_selecao", item.get("_score", 0))
            s_final   = item.get("_score", 0)
            exp = item.get("_explanation", "")
            rank = item.get("_rank")

            def _extract(prefix_str):
                parts = [p for p in exp.split(" | ") if prefix_str in p]
                return "; ".join(parts)

            rep_penalty = s_final - s_selecao  # 0 ou negativo

            audit_rows.append({
                "source": name, "title": item["title"], "url": item["url"],
                "horario": format_time(item.get("dt")),
                "topico":      _article_topic(item["url"]) or "",
                # Nota de seleção: score usado para entrar no top 20 (comparável entre todos)
                "score":       s_selecao,
                "rank":        rank if rank else "",
                "s_acao":      _extract("ação"),
                "s_impacto":   _extract("impacto"),
                "s_especif":   _extract("especif") or _extract("percentual") or _extract("numérico") or _extract("manchete obj") or _extract("prazo"),
                "s_tendencia": _extract("tendência"),
                "s_editoria":  _extract("editoria"),
                "s_penalidade":_extract("-"),
                "s_repeticao": str(rep_penalty) if rep_penalty < 0 else "",
                "status":      "Top 20 — disponível" if rank else "Score baixo / não entrou",
                "explicacao":  exp,
                "_item_ref":   item,
            })

        print(f"  → {len(top)} no top 20 / {len(in_win)} na janela / {len(out_of_win)} fora / {len(no_date)} sem data")

    # ── Fase 4: exibe manchetes ────────────────────────────────────
    print("\n\n" + "═" * 55)
    print("  MANCHETES SELECIONADAS")
    print("═" * 55)

    for name, config in SOURCES.items():
        items = all_selected.get(name, [])
        if items:
            print_section(name, config["prefix"], items)
        else:
            print(f"\n[{name}] — nenhuma manchete encontrada")

    # ── Fase 5: seleção do usuário ─────────────────────────────────
    print("\n" + "═" * 55)
    print("\nDigite os códigos das manchetes (ex: V1 V3 E2 E5 G1 G4 F2 F6 — até 20 por veículo):\n")
    raw = input("→ ").strip().upper()
    if not raw:
        print("Nenhum código digitado. Encerrando.")
        return

    grouped: Dict[str, List[dict]] = {name: [] for name in SOURCES}
    unknown = []
    for code in raw.split():
        if code in code_map:
            item = code_map[code]
            grouped[item["source"]].append(item)
        else:
            unknown.append(code)

    if unknown:
        print(f"\n[aviso] Códigos não encontrados: {', '.join(unknown)}")

    # Atualiza status dos selecionados
    selected_urls = {i["url"] for lst in grouped.values() for i in lst}
    for row in audit_rows:
        if row.get("_item_ref", {}).get("url") in selected_urls:
            row["status"] = "Selecionado"
    for row in audit_rows:
        row.pop("_item_ref", None)
    # Ordena planilha: selecionados primeiro, depois top 20, depois excluídos
    status_order = {"Selecionado": 0, "Top 20 — disponível": 1,
                    "Score baixo / não entrou": 2, "Fora da janela de tempo": 3,
                    "Sem data identificável": 4}
    audit_rows.sort(key=lambda r: (
        list(SOURCES.keys()).index(r["source"]) if r["source"] in SOURCES else 99,
        status_order.get(r.get("status", ""), 5),
        -int(r["score"]) if isinstance(r.get("score"), int) else 0,
    ))

    # ── Fase 6: clipping ──────────────────────────────────────────
    today_str = datetime.now(tz=LOCAL_TZ).strftime("%d/%m/%Y")
    lines = [
        "*Necton Markets | Clipping de notícias*", "",
        f"*{today_str}*", "",
        "_*Confira as principais notícias dos jornais mais relevantes do país*_", "",
    ]
    for name in SOURCES:
        items = grouped[name]
        if not items:
            continue
        lines.append(f"*{name}*")
        lines.append("")
        for item in items:
            short_url = _shorten_url(item['url'])
            lines.append(f"- {item['title']} {short_url}")
        lines.append("")
    lines.append("*Necton Investimentos*")
    clipping = "\n".join(lines)

    print("\n" + "═" * 55)
    print("  CLIPPING FINAL")
    print("═" * 55 + "\n")
    print(clipping)
    print("\n" + "═" * 55)

    import os
    clipping_path = os.path.join(os.path.expanduser("~"), "Desktop", "clipping_final.txt")
    with open(clipping_path, "w", encoding="utf-8") as f:
        f.write(clipping + "\n")
    print(f"\n✅ Clipping salvo em {clipping_path}")

    generate_spreadsheet(audit_rows, cluster_stats)


if __name__ == "__main__":
    import sys as _sys
    if len(_sys.argv) > 1 and _sys.argv[1] == "--retool":
        # Modo Retool RPC: fica escutando chamadas do Retool
        try:
            import retoolrpc
            rpc = retoolrpc.RetoolRPC(
                api_key=_sys.argv[2] if len(_sys.argv) > 2 else "",
                host="https://api.retool.com/rpc",
            )
            rpc.register(retool_buscar_manchetes)
            rpc.register(retool_gerar_clipping)
            rpc.register(retool_gerar_planilha)
            print("✅ Retool RPC ativo. Aguardando chamadas...")
            rpc.listen()
        except ImportError:
            print("❌ retoolrpc não instalado. Execute: pip install retoolrpc")
    else:
        main()
